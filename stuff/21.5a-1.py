import tkinter as tk
from tkinter import simpledialog
from datetime import datetime
import psutil
import RPi.GPIO as GPIO
import time
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import load_workbook
#pip3 install openpyxl sudo apt install python3-openpyxl

# ביטול אזהרות
GPIO.setwarnings(False)

# הגדרת פינים
valve_pins = [2, 3, 14, 15, 10, 11]  # פינים לשסתומים
limit_switch_pins = [24, 23, 22, 18, 17, 4, 7, 8, 27, 25]  # פינים למפסיקי גבול
S0, S1, S2, S3 = 6, 12, 13, 16  # פינים לבחירת ערוץ במולטיפלקסור
SIG_PIN = 5  # פין היציאה מהמולטיפלקסור

# הגדרת פינים
motor_pins = {
    "X": {"step": 19, "dir": 20},  # מנוע 1: פין צעד ופין כיוון
    "Y": {"step": 21, "dir": 26}  # מנוע 2: פין צעד ופין כיוון
}

global row,column,EXCEL_FILE_PATH,num_step_x,max_mm_x

row = 3
column = 4
start_delay=0.01
homing_delay=0.1
max_mm_x=10
num_step_x=0
max_mm_y=10
num_step_y=0
step_per_mm_x=7
step_per_mm_y=7

# נתיב קובץ האקסל
EXCEL_FILE_PATH ="/home/ori/Documents/Data1.xlsx"#/media/ori/000C-FA3C/Data1.xlsx"

# אתחול GPIO
GPIO.setmode(GPIO.BCM)
for motor in motor_pins.values():
    GPIO.setup(motor["step"], GPIO.OUT)
    GPIO.setup(motor["dir"], GPIO.OUT)
    
# הגדרת פינים לשסתומים
for pin in valve_pins:
    GPIO.setup(pin, GPIO.OUT)
    GPIO.output(pin, GPIO.LOW)
# הגדרת פינים למפסיקי גבול
for pin in limit_switch_pins:
    GPIO.setup(pin, GPIO.IN)

# הגדרת פינים לבחירת ערוץ במולטיפלקסור
GPIO.setup(S0, GPIO.OUT)
GPIO.setup(S1, GPIO.OUT)
GPIO.setup(S2, GPIO.OUT)
GPIO.setup(S3, GPIO.OUT)

# הגדרת פין היציאה למולטיפלקסור
GPIO.setup(SIG_PIN, GPIO.IN)


# Create the main window
root = tk.Tk()
root.title("User Interface with Frames")
root.geometry("2000x1200")  # Initial window size
#zmni                  root.attributes('-fullscreen', True)
#root.bind("<Escape>", lambda event: root.attributes('-fullscreen', False))

# Variables for user input and movement amount
selected_program = tk.StringVar()
movement_up_amount = tk.StringVar()
movement_down_amount = tk.StringVar()
movement_right_amount= tk.StringVar()
movement_left_amount= tk.StringVar()
relative_position_y=tk.StringVar()
relative_position_y.set("9999")
abslute_position_y=tk.StringVar()
abslute_position_y.set('9999')
relative_position_x=tk.StringVar()
relative_position_x.set('9999')
abslute_position_x=tk.StringVar()
abslute_position_x.set('9999')

















    
def move_motor(motor_id, direction, mm_movment, step_delay):
    # בדיקת מזהה מנוע
    if motor_id not in motor_pins:
        print(f"Error: Motor {motor_id} is not defined.")
        return

    # חישוב צעדים לפי יחס המרה
    if motor_id == "X":
        steps = int(mm_movment * step_per_mm_x)
    elif motor_id == "Y":
        steps = int(mm_movment * step_per_mm_y)
    else:
        print(f"Error: No step-per-mm definition for motor {motor_id}.")
        return

    motor = motor_pins[motor_id]
    GPIO.output(motor["dir"], direction)

    for _ in range(steps):
        GPIO.output(motor["step"], GPIO.HIGH)
        time.sleep(step_delay)
        GPIO.output(motor["step"], GPIO.LOW)
        time.sleep(step_delay)

    print(f"Motor {motor_id} moved {steps} steps in {'forward' if direction == 1 else 'backward'} direction.")






def a_function(num):# תנועת שרטוט מטה
    global row,column
    print('A',num)
    if read_sensor(1) == 1:
        print("חיישן בערוץ 8 הופעל")
    toggle_valve(1)
    move_motor('X', 10, num,start_delay)
    print (column)
    column=column+1
    print (column)
    activate()  

def b_function(num):#תנועת שרטוט מעלה
    global row,column    
    print('B',num)
    
    if GPIO.input(22) == GPIO.HIGH:
        print("❗ לא ניתן לנוע מעלה - מפסק גבול פעיל!")
        return  # יוצאים מהפונקציה בלי להפעיל את המנוע
    
    move_motor('X', 0, num,start_delay)
    print (column)
    column=column+1
    print (column)
    activate()
    
def c_function(num):#סיום שרטוט ותנוע ל?
    print('C',num)
    move_motor('X', 1, num,start_delay)
    
def d_function(num):
    print('D',num)

    

def e_function(num):#תנועת חלוקה ימינה
    print('E',num)
    move_motor(Y, 1, num,start_delay)
    
def f_function(num):#תנועת חלוקה שמאלה
    print('F',num)
    move_motor(Y, 0, num,start_delay)

def g_function(num):#סיום חלוקה ותנועה שמאלה
    print('G',num)
    move_motor(Y, 0, num,start_delay)
    

def h_function(num):
    print('H',num)
    
    
def i_function(num):
    print('I',num)
    
   
   
   
   
   
   
   
   
   
# הפונקציה מקבלת את הערך שבתא ומחזירה את הפונקציה שצריך להפעיל ואת המספר להפעלה   
def execute_command(command):
    # Extract the first character (the letter) and the rest (the number)
    letter = command[0]# התו הראשון נשמר
    number = int(command[1:])# מהתו השני עד הסןף נשמר כמספר
    # Build the function name
    func_name = f"{letter.lower()}_function"
    # Get the function from the global namespace and call it
    if func_name in globals():#מוודאים שהפונקציה קיימת
        func = globals()[func_name]
        return func(number)# מחזירה את הפנוקציה להפעלה עם המספר
    else:
        return f"Function '{func_name}' not found!"# אם הפונקציה לא קיימת מחזירה "לא נמצאה"
#def execute_command(command):
#     print('exe')
#     if len(command) < 2 or not command[1:].isdigit():
#         return "שגיאה: הפקודה צריכה להכיל אות ואחריה מספר, למשל 'A5'"
#     
#     letter = command[0]
#     number = int(command[1:])
#     func_name = f"{letter.lower()}_function"
#     if func_name in globals():
#         func = globals()[func_name]
#         return func(number)
#     else:
#         return f"Function '{func_name}' not found!"

# Function to activate the selected program
def activate():# מופעלת בלחיצה על כפתור ההפעלה ומחזירה את הערך בתא שהאינדקסים מצביעים עליו
    #print('activ')
    number = selected_program.get()# לוקחת את הערך שהמשתמש הכניס - איזה תוכנית להפעיל
    row=int(number)# הופכים את הערך מתו למספר
    print(f"Selected number: {number},233")
    # Additional logic can be added here
    try:
        # טוען את קובץ האקסל
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH) # פתיחת קובץ האקסל ע"פ הנתיב
        sheet = workbook.active  # משתמש בגליון הראשון הפעיל 
        # מחזיר את הערך בתא
        value= sheet.cell(row=row, column=column).value # הוצאת הערך מתוך התא שהאינדקסים משוייכים אליו
        print (value,'241')
        execute_command(value) #קריאה לפוקציה שמפרשת את הערך בתוך התא
    except FileNotFoundError:
        print(f"Error: File {EXCEL_FILE_PATH} not found.,244")
    except Exception as e:
        print(f"An error occurred: {e},246")

def homing():    #פונקציה לאיפוס הצירים
    if GPIO.input(27) == GPIO.LOW:#and GPIO.input(27) == GPIO.HIGH:#אם מפסק הגבול השמאלי של החלוקה לא לחוץ
        print("(27) == GPIO.LOW")
        num_step_x=0 # איפוס מונה צעדים - נועד להגביל את מספר הצעדים האפשריים
        while GPIO.input(27) == GPIO.LOW and max_mm_x> num_step_x:
            move_motor('X', 1, 1,homing_delay)
            num_step_x=num_step_x+1
    if GPIO.input(27) == GPIO.HIGH:#and GPIO.input(27) == GPIO.HIGH:#אם המפסק גבול נלחץ מאפסים את המיקומים המוחלט והיחסי
        relative_position_x.set("0")
        abslute_position_x.set("0")
    if GPIO.input(25) == GPIO.LOW:#and GPIO.input(25) == GPIO.HIGH:
        print("(25) == GPIO.LOW")
        num_step_y=0
        while GPIO.input(25) == GPIO.LOW and max_mm_y> num_step_y:
            move_motor('Y', 1, 1,homing_delay)
            num_step_y=num_step_y+1
    if GPIO.input(25) == GPIO.HIGH:
        relative_position_y.set("0")
        abslute_position_y.set("0")


    # Additional logic can be added here
def exit_fullscreen():
    # יצירת חלון להזנת סיסמה
    password = simpledialog.askstring("אימות סיסמה", "הזן את הסיסמה ליציאה ממסך מלא:", show="*")
    if password == "1234":  # החליפי לדרישת הסיסמה שלך
        root.attributes('-fullscreen', False)
    else:
        tk.messagebox.showerror("שגיאה", "סיסמה שגויה! נסה שוב.")
        
# פונקציה לגלילת עכבר מזיז את התצוגה עם הגלילת עכבר
def on_mouse_scroll(event):
    canvas.yview_scroll(-1 * int(event.delta / 120), "units")

# פונקציה לטעינת קובץ Excel הופכת את הקובץ לפעיל
def load_excel():
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
        display_data(sheet)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file:\n{e}")
#פונקציה שמעבירה את המידע מקובץ האקסל לטבלת TREEVIEW שזמינה לTKINTER
def display_data(sheet):
    for widget in table_frame.winfo_children():
        widget.destroy()
    # Create a Treeview widget
    tree = ttk.Treeview(table_frame, show="headings")
    tree.place(relx=0, rely=0, relwidth=1, relheight=0.95)  # Position and size (60% of the window)
    # Add horizontal scrollbar
    h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
    h_scrollbar.place(relx=0, rely=0.9, relwidth=1, relheight=0.05)  # Place below the table
    tree.configure(xscrollcommand=h_scrollbar.set)

    # Set up columns
    columns = [cell.value for cell in sheet[1]]  # Assume first row is headers
    if not columns:
        messagebox.showinfo("Info", "The Excel file is empty.")
        return

    tree["columns"] = columns
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=100)

    # Insert rows, skipping empty rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):  # Skip empty rows
            tree.insert("", tk.END, values=row)



# פונקציה לבחירת ערוץ במולטיפלקסור- מקבלת מספר ערוץ להפעלה ומדליקה את היציאות בהתאם
def select_channel(channel):
    GPIO.output(S0, channel & 0x01)
    GPIO.output(S1, (channel >> 1) & 0x01)
    GPIO.output(S2, (channel >> 2) & 0x01)
    GPIO.output(S3, (channel >> 3) & 0x01)
    time.sleep(0.01)  # השהיה קצרה לאחר החלפת הערוץ

# פונקציה לקריאת חיישן דרך המולטיפלקסור- מקבלת ערוץ לקריאה קוראת לפונקציה שמפעילה את הערוץ ואז קוראת את התוצאה ומחזירה את הערך שלו
def read_sensor(channel):
    select_channel(channel)
    return GPIO.input(SIG_PIN)

# פונקציה להפעלת שסתום- מקבלת מספר שסתום והופכת את מצבו , קוראת לפונקציה לשינוי מצב התצוגה
def toggle_valve(valve_num):
    current_state = GPIO.input(valve_pins[valve_num])
    GPIO.output(valve_pins[valve_num], not current_state)
    update_button_text(valve_num)


# פונקציה לעדכון טקסט הכפתור בהתאם למצב השסתום
def update_button_text(valve_num):
    if GPIO.input(valve_pins[valve_num]) == GPIO.HIGH:
        buttons[valve_num].config(text=f"Turn off Valve {valve_num + 1}")
    else:
        buttons[valve_num].config(text=f"Turn on Valve {valve_num + 1}")

# פונקציה לעדכון התצוגה של חיווי החיישנים
def update_sensor_indicators():
    for i in range(16):
        sensor_value = read_sensor(i)
        if sensor_value == GPIO.HIGH:
            sensor_labels[i].config(bg="green", text=f"Sen{i + 1}: ON")
        else:
            sensor_labels[i].config(bg="red", text=f"Sen{i + 1}: OFF")
    root.after(500, update_sensor_indicators)  # עדכון כל 500 מילישניות

# פונקציה לעדכון התצוגה של חיווי מפסיקי הגבול
def update_limit_switch_indicators():
    for i in range(10):
        if GPIO.input(limit_switch_pins[i]) == GPIO.HIGH:
            limit_switch_labels[i].config(bg="green", text=f"Limit Switch {i + 1}: ON")
        else:
            limit_switch_labels[i].config(bg="red", text=f"Limit Switch {i + 1}: OFF")
    root.after(500, update_limit_switch_indicators)  # עדכון כל 500 מילישניות


          
# Function to create content פונקציה לייצירת תוכנית חדשה- התמעת שורה באקסל
def create_content():
    print("Function 'Create Content' activated",'370')   
    
def previos_row():
    print("Function 'previos row' activated",)   
def Move_Up_to_End():
    print("Function 'Move_Up_to_End' activated")    
def Special_Move_Up():
    program_num=movement_up_amount.get()
    print("Function 'Special_Move_Up' activated", program_num )
def Manual_Move_Up():
    print("Function 'Manual_Move_Up' activated")
def Move_Up_10_mm():
    print("Function 'Move_Up_10_mm' activated")
    move_motor('Y', 0, 10,start_delay)
def Move_Up_5_mm():
    print("Function 'Move_Up_5_mm' activated")
def Move_Up_1_mm():
    print("Function 'Move_Up_1_mm' activated")
def Move_Up_05_mm():
    print("Function 'Move_Up_05_mm' activated")
def Move_down_to_End():
    print("Function 'Move_down_to_End' activated")    
def Special_Move_down():
    program_num=movement_down_amount.get()
    print("Function 'Special_Move_down' activated", program_num )
def Manual_Move_down():
    print("Function 'Manual_Move_down' activated")
def Move_down_10_mm():
    print("Function 'Move_down_10_mm' activated")
    move_motor('Y', 1, 10,start_delay)
def Move_down_5_mm():
    print("Function 'Move_down_5_mm' activated")
def Move_down_1_mm():
    print("Function 'Move_down_1_mm' activated")
def Move_down_05_mm():
    print("Function 'Move_down_05_mm' activated")
def next_line():
    print("Function 'next line' activated")
def next_column():
    print("Function 'next column' activated")
def Move_Right_to_End():
    print("Function 'Move_Right_to_End' activated")    
def Special_Move_Right():
    program_num = movement_right_amount.get()  # שים לב לעדכון כאן
    print(f"Function 'Special_Move_Right' activated",program_num) 
def Manual_Move_Right():
    print("Function 'Manual_Move_Right' activated")
def Move_Right_10_mm():
    print("Function 'Move_Right_10_mm' activated")
def Move_Right_5_mm():
    print("Function 'Move_Right_5_mm' activated")
def Move_Right_1_mm():
    print("Function 'Move_Right_1_mm' activated")
def Move_Right_05_mm():
    print("Function 'Move_Right_05_mm' activated")
def Move_Left_to_End():
    print("Function 'Move_Left_to_End' activated")    
def Special_Move_Left():
    program_num = movement_left_amount.get()  # עדכון המשתנה כאן
    print(f"Function 'Special_Move_Left' activated", program_num)
def Manual_Move_Left():
    print("Function 'Manual_Move_Left' activated")
def Move_Left_10_mm():
    print("Function 'Move_Left_10_mm' activated")
def Move_Left_5_mm():
    print("Function 'Move_Left_5_mm' activated")  
def Move_Left_1_mm():
    print("Function 'Move_Left_1_mm' activated")
def Move_Left_05_mm():
    print("Function 'Move_Left_05_mm' activated")
def previos_column():
    print("Function 'previos_column' activated")
def reset_relative_position_y():
    print("Function 'reset_relative_position_y' activated")
    relative_position_y.set("0")
def reset_relative_position_x():
    print("Function 'reset_relative_position_x' activated")
    relative_position_x.set("0")
    
    
    
    
    
    
    
# Function to update the label with time, date, and CPU temperature
def update_variables():
    current_time = datetime.now().strftime("%H:%M:%S")
    current_date = datetime.now().strftime("%Y-%m-%d")
    cpu_temp = psutil.sensors_temperatures().get('cpu-thermal', [{}])[0].get('current', 'N/A')
    #print(cpu_temp)
    label_vars.config(text=f"Time: {current_time}, Date: {current_date}, CPU Temp: {cpu_temp}°C")
    root.after(1000, update_variables)







################################################################################
# מסגרת עליונה
top_frame = tk.Frame(root, bg="lightblue")
top_frame.place(relx=0, rely=0, relwidth=1, relheight=0.1)


# כפתור ליציאה ממסך מלא עם סיסמה
exit_button = tk.Button(top_frame, text="יציאה ממסך מלא", command=exit_fullscreen)
exit_button.place(relx=0, rely=0, relwidth=0.15, relheight=1)

# Frame around the entry box
frame_entry = tk.LabelFrame(top_frame, text="Select Program בחר תוכנית", padx=5, pady=5)
frame_entry.place(relx=0.15, rely=0, relwidth=0.15, relheight=1)

# Entry box inside the frame
entry_program = tk.Entry(frame_entry, textvariable=selected_program)
entry_program.pack()

# show button
button_show = tk.Button(top_frame, text="homing איפוס", command=homing)
button_show.place(relx=0.3, rely=0, relwidth=0.15, relheight=1)

# Activate button
button_run = tk.Button(top_frame, text="Activate הפעל", command=activate)
button_run.place(relx=.45, rely=0, relwidth=0.15, relheight=1)

# "Create Content" button
button_create = tk.Button(top_frame, text="Create Content צור תוכנית", command=create_content)
button_create.place(relx=0.6, rely=0, relwidth=0.15, relheight=1)

# Label displaying variables (time, date, CPU temperature)
label_vars = tk.Label(top_frame, text="Time: --:--:--, Date: ----/--/--, CPU Temp: --°C")
label_vars.place(relx=0.75, rely=0, relwidth=0.25, relheight=1)

# Initial call to update the label
#update_variables()
















#############################################################
# מסגרת ימנית
right_frame = tk.Frame(root)
right_frame.place(relx=0.9, rely=0.1, relwidth=0.1, relheight=0.95)
right_frame_relheight=0.041

# Adding buttons and entry fields in the right frame

button_previos_row = tk.Button(right_frame, text="Move Up to previos row", command=previos_row)
button_previos_row.place(relx=0, rely=0, relwidth=1, relheight=right_frame_relheight)

button_up_end = tk.Button(right_frame, text="Move Up to End", command=Move_Up_to_End)
button_up_end.place(relx=0, rely=(right_frame_relheight*1), relwidth=1, relheight=right_frame_relheight)

button_special_up = tk.Button(right_frame, text="Special Move Up", command=Special_Move_Up)
button_special_up.place(relx=0, rely=(right_frame_relheight*2), relwidth=1, relheight=right_frame_relheight)

entry_movement_amount = tk.Entry(right_frame, textvariable=movement_up_amount)
entry_movement_amount.place(relx=0, rely=(right_frame_relheight*3), relwidth=1, relheight=right_frame_relheight)

button_manual_up = tk.Button(right_frame, text="Manual Move Up", command=Manual_Move_Up)
button_manual_up.place(relx=0, rely=(right_frame_relheight*4), relwidth=1, relheight=right_frame_relheight)

button_up_10mm = tk.Button(right_frame, text="Move Up 10 mm", command=Move_Up_10_mm)
button_up_10mm.place(relx=0, rely=(right_frame_relheight*5), relwidth=1, relheight=right_frame_relheight)

button_manual_up = tk.Button(right_frame, text="Move Up 5 mm", command=Move_Up_5_mm)
button_manual_up.place(relx=0, rely=(right_frame_relheight*6), relwidth=1, relheight=right_frame_relheight)

button_up_10mm = tk.Button(right_frame, text="Move Up 1 mm", command=Move_Up_1_mm)
button_up_10mm.place(relx=0, rely=(right_frame_relheight*7), relwidth=1, relheight=right_frame_relheight)

button_up_10mm = tk.Button(right_frame, text="Move Up 1/2 mm", command=Move_Up_05_mm)
button_up_10mm.place(relx=0, rely=(right_frame_relheight*8), relwidth=1, relheight=right_frame_relheight)

# Frame around the relative_position_y
frame_relative_position_y = tk.LabelFrame(right_frame, text="relative_position_y ", padx=5, pady=5)
frame_relative_position_y.place(relx=0, rely=(right_frame_relheight*9), relwidth=1, relheight=right_frame_relheight*2)

label_relative_position_y = tk.Label(frame_relative_position_y, textvariable=relative_position_y)
label_relative_position_y.place(relx=0, rely=0, relwidth=1, relheight=1)

button_reset_relative_position_y = tk.Button(right_frame, text="reset relative position y", command=reset_relative_position_y)
button_reset_relative_position_y.place(relx=0, rely=(right_frame_relheight*11), relwidth=1, relheight=right_frame_relheight)

# Frame around the abslute_position_y
frame_abslute_position_y = tk.LabelFrame(right_frame, text="abslute_position_y ", padx=5, pady=5)
frame_abslute_position_y.place(relx=0, rely=(right_frame_relheight*12), relwidth=1, relheight=right_frame_relheight*2)

label_abslute_position_y = tk.Label(frame_abslute_position_y, textvariable=abslute_position_y)
label_abslute_position_y.place(relx=0, rely=0, relwidth=1, relheight=1)

button_down_end = tk.Button(right_frame, text="Move down to End", command=Move_down_to_End)
button_down_end.place(relx=0, rely=(right_frame_relheight*14), relwidth=1, relheight=right_frame_relheight)

button_special_down = tk.Button(right_frame, text="Special Move down", command=Special_Move_down)
button_special_down.place(relx=0, rely=(right_frame_relheight*15), relwidth=1, relheight=right_frame_relheight)

entry_movement_amount = tk.Entry(right_frame, textvariable=movement_down_amount)
entry_movement_amount.place(relx=0, rely=(right_frame_relheight*16), relwidth=1, relheight=right_frame_relheight)

button_manual_down = tk.Button(right_frame, text="Manual Move down", command=Manual_Move_down)
button_manual_down.place(relx=0, rely=(right_frame_relheight*17), relwidth=1, relheight=right_frame_relheight)

button_down_10mm = tk.Button(right_frame, text="Move down 10 mm", command=Move_down_10_mm)
button_down_10mm.place(relx=0, rely=(right_frame_relheight*18), relwidth=1, relheight=right_frame_relheight)

button_manual_down = tk.Button(right_frame, text="Move down 5 mm", command=Move_down_5_mm)
button_manual_down.place(relx=0, rely=(right_frame_relheight*19), relwidth=1, relheight=right_frame_relheight)

button_down_1mm = tk.Button(right_frame, text="Move down 1 mm", command=Move_down_1_mm)
button_down_1mm.place(relx=0, rely=(right_frame_relheight*20), relwidth=1, relheight=right_frame_relheight)

button_down_05mm = tk.Button(right_frame, text="Move down 1/2 mm", command=Move_down_05_mm)
button_down_05mm.place(relx=0, rely=(right_frame_relheight*21), relwidth=1, relheight=right_frame_relheight)

button_next_line = tk.Button(right_frame, text="Move down to next row", command=next_line)
button_next_line.place(relx=0, rely=(right_frame_relheight*22), relwidth=1, relheight=right_frame_relheight)


###################################################

# מסגרת ימנית מרכזית
center_right_frame = tk.Frame(root)
center_right_frame.place(relx=0.8, rely=0.1, relwidth=0.1, relheight=0.95)

next_column = tk.Button(center_right_frame, text="Move Right to next column", command=next_column)
next_column.place(relx=0, rely=(right_frame_relheight*0), relwidth=1, relheight=right_frame_relheight)

button_right_end = tk.Button(center_right_frame, text="Move Right to End", command=Move_Right_to_End)
button_right_end.place(relx=0, rely=(right_frame_relheight*1), relwidth=1, relheight=right_frame_relheight)

button_special_right = tk.Button(center_right_frame, text="Special Move Right", command=Special_Move_Right)
button_special_right.place(relx=0, rely=(right_frame_relheight*2), relwidth=1, relheight=right_frame_relheight)

entry_movement_amount_right = tk.Entry(center_right_frame, textvariable=movement_right_amount)
entry_movement_amount_right.place(relx=0, rely=(right_frame_relheight*3), relwidth=1, relheight=right_frame_relheight)

button_manual_right = tk.Button(center_right_frame, text="Manual Move Right", command=Manual_Move_Right)
button_manual_right.place(relx=0, rely=(right_frame_relheight*4), relwidth=1, relheight=right_frame_relheight)

button_right_10mm = tk.Button(center_right_frame, text="Move Right 10 mm", command=Move_Right_10_mm)
button_right_10mm.place(relx=0, rely=(right_frame_relheight*5), relwidth=1, relheight=right_frame_relheight)

button_manual_right = tk.Button(center_right_frame, text="Move Right 5 mm", command=Move_Right_5_mm)
button_manual_right.place(relx=0, rely=(right_frame_relheight*6), relwidth=1, relheight=right_frame_relheight)

button_right_10mm = tk.Button(center_right_frame, text="Move Right 1 mm", command=Move_Right_1_mm)
button_right_10mm.place(relx=0, rely=(right_frame_relheight*7), relwidth=1, relheight=right_frame_relheight)

button_right_10mm = tk.Button(center_right_frame, text="Move Right 1/2 mm", command=Move_Right_05_mm)
button_right_10mm.place(relx=0, rely=(right_frame_relheight*8), relwidth=1, relheight=right_frame_relheight)

# Frame around the relative_position_x
frame_relative_position_x = tk.LabelFrame(center_right_frame, text="relative_position_x ", padx=5, pady=5)
frame_relative_position_x.place(relx=0, rely=(right_frame_relheight*9), relwidth=1, relheight=right_frame_relheight*2)

label_relative_position_x = tk.Label(frame_relative_position_x, textvariable=relative_position_x)
label_relative_position_x.place(relx=0, rely=0, relwidth=1, relheight=1)

button_reset_relative_position_x = tk.Button(center_right_frame, text="reset relative position x", command=reset_relative_position_x)
button_reset_relative_position_x.place(relx=0, rely=(right_frame_relheight*11), relwidth=1, relheight=right_frame_relheight)

# Frame around the abslute_position_x
frame_abslute_position_x = tk.LabelFrame(center_right_frame, text="abslute_position_x ", padx=5, pady=5)
frame_abslute_position_x.place(relx=0, rely=(right_frame_relheight*12), relwidth=1, relheight=right_frame_relheight*2)

label_abslute_position_x = tk.Label(frame_abslute_position_x, textvariable=abslute_position_x)
label_abslute_position_x.place(relx=0, rely=0, relwidth=1, relheight=1)

button_left_end = tk.Button(center_right_frame, text="Move Left to End", command=Move_Left_to_End)
button_left_end.place(relx=0, rely=(right_frame_relheight*14), relwidth=1, relheight=right_frame_relheight)

button_special_left = tk.Button(center_right_frame, text="Special Move Left", command=Special_Move_Left)
button_special_left.place(relx=0, rely=(right_frame_relheight*15), relwidth=1, relheight=right_frame_relheight)

entry_movement_amount_left = tk.Entry(center_right_frame, textvariable=movement_left_amount)
entry_movement_amount_left.place(relx=0, rely=(right_frame_relheight*16), relwidth=1, relheight=right_frame_relheight)

button_manual_left = tk.Button(center_right_frame, text="Manual Move Left", command=Manual_Move_Left)
button_manual_left.place(relx=0, rely=(right_frame_relheight*17), relwidth=1, relheight=right_frame_relheight)

button_left_10mm = tk.Button(center_right_frame, text="Move Left 10 mm", command=Move_Left_10_mm)
button_left_10mm.place(relx=0, rely=(right_frame_relheight*18), relwidth=1, relheight=right_frame_relheight)

button_manual_left = tk.Button(center_right_frame, text="Move Left 5 mm", command=Move_Left_5_mm)
button_manual_left.place(relx=0, rely=(right_frame_relheight*19), relwidth=1, relheight=right_frame_relheight)

button_left_10mm = tk.Button(center_right_frame, text="Move Left 1 mm", command=Move_Left_1_mm)
button_left_10mm.place(relx=0, rely=(right_frame_relheight*20), relwidth=1, relheight=right_frame_relheight)

button_left_10mm = tk.Button(center_right_frame, text="Move Left 1/2 mm", command=Move_Left_05_mm)
button_left_10mm.place(relx=0, rely=(right_frame_relheight*21), relwidth=1, relheight=right_frame_relheight)

previos_column = tk.Button(center_right_frame, text="Move left to previos column", command=previos_column)
previos_column.place(relx=0, rely=(right_frame_relheight*22), relwidth=1, relheight=right_frame_relheight)












































##############################################################################
# מסגרת שמאלית
left_frame = tk.Frame(root, bg="green")
left_frame.place(relx=0, rely=0.1, relwidth=0.1, relheight=0.95)

left_frame_relheight=0.0289


# כפתורי שסתומים
buttons = []
for i in range(6):
    button = tk.Button(left_frame, text=f"Turn on Valve {i + 1}", command=lambda i=i: toggle_valve(i), width=10, height=1)
    #button.grid(row=i, column=0, padx=10, pady=2)
    button.place(relx=0, rely=(left_frame_relheight*i), relwidth=1, relheight=left_frame_relheight)
    buttons.append(button)



# תוויות חיווי לחיישנים
sensor_labels = []
for i in range(16):
    label = tk.Label(left_frame, text=f"Sensor {i + 1}: OFF", width=10, height=1, bg="red", fg="white")
    #label.grid(row=i+6, column=0, padx=10, pady=2)
    label.place(relx=0, rely=(left_frame_relheight*(i+6)), relwidth=1, relheight=left_frame_relheight)

    sensor_labels.append(label)

# תוויות חיווי למפסיקי גבול
limit_switch_labels = []
for i in range(10):
    label = tk.Label(left_frame, text=f"Limit Switch {i + 1}: OFF", width=10, height=1, bg="red", fg="white")
    #label.grid(row=i+6+16, column=0, padx=10, pady=2)
    label.place(relx=0, rely=(left_frame_relheight*(i+6+16)), relwidth=1, relheight=left_frame_relheight)

    limit_switch_labels.append(label)

# עדכון כפתורי השסתומים במצב התחלתי
for i in range(6):
    update_button_text(i)

# התחלת עדכון חיווי החיישנים ומפסיקי הגבול
root.after(500, update_sensor_indicators)
root.after(500, update_limit_switch_indicators)






#################################################################################################
# מסגרת שמאלית מרכזית
center_left_frame = tk.Frame(root, bg="gray")
center_left_frame.place(relx=0.1, rely=0.1, relwidth=0.7, relheight=0.95)

# Frame for table
table_frame = tk.Frame(center_left_frame)
table_frame.pack(fill=tk.BOTH, expand=True)

# Load Excel file at startup
load_excel()



# סגירה בטוחה של GPIO כאשר החלון נסגר
def on_closing():
    GPIO.cleanup()
    root.destroy()

root.protocol("WM_DELETE_WINDOW", on_closing)
root.mainloop()










